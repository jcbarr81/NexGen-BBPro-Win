#!/usr/bin/env python3
"""Compute per-team injury baselines from the roster-resource report.

Reads ``data/MLB_avg/roster-resource__injury-report.xlsx`` and aggregates
injury counts along with days missed (when both an injury date and return
date are provided). The output helps scale the in-game injury simulator to
leagues with fewer than 30 MLB teams.
"""

from __future__ import annotations

from collections import defaultdict, Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, Optional
import argparse
import json

from openpyxl import load_workbook

DEFAULT_WORKBOOK = Path("data/MLB_avg/roster-resource__injury-report.xlsx")
DATE_FORMATS = ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%b %d %Y"]


def _coerce_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serialized date -- openpyxl usually converts already, but guard.
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value)).date()
        except ValueError:
            return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                continue
    return None


@dataclass
class InjuryRecord:
    team: str
    position: str
    injury_date: Optional[date]
    return_date: Optional[date]

    @property
    def days_missed(self) -> Optional[int]:
        if self.injury_date and self.return_date:
            delta = (self.return_date - self.injury_date).days
            if delta >= 0:
                return delta
        return None

    @property
    def is_pitcher(self) -> bool:
        pos = (self.position or "").upper()
        if not pos:
            return False
        return pos in {"P", "SP", "RP"} or pos.startswith("P/")


def load_injury_records(workbook_path: Path) -> Iterable[InjuryRecord]:
    wb = load_workbook(workbook_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            team = row[1] or sheet_name
            pos = row[2] or ""
            inj_date = _coerce_date(row[3])
            return_date = _coerce_date(row[8]) or _coerce_date(row[7])
            if not any((name, inj_date, return_date)):
                continue
            yield InjuryRecord(team=str(team), position=str(pos), injury_date=inj_date, return_date=return_date)


def summarize(records: Iterable[InjuryRecord]) -> Dict[str, object]:
    counts = Counter()
    day_totals = defaultdict(int)
    tracked_days = 0
    day_samples = 0
    pitcher_count = 0
    total_records = 0

    for rec in records:
        total_records += 1
        counts[rec.team] += 1
        if rec.days_missed is not None:
            day_totals[rec.team] += rec.days_missed
            tracked_days += rec.days_missed
            day_samples += 1
        if rec.is_pitcher:
            pitcher_count += 1

    team_count = len(counts) or 1
    total_injuries = sum(counts.values())
    avg_injuries_per_team = total_injuries / team_count
    avg_days_per_team = (sum(day_totals.values()) / team_count) if day_totals else 0.0
    avg_days_per_stint = (tracked_days / day_samples) if day_samples else 0.0
    pitcher_share = pitcher_count / total_records if total_records else 0.0

    return {
        "teams_tracked": team_count,
        "total_records": total_records,
        "avg_injuries_per_team": avg_injuries_per_team,
        "avg_days_lost_per_team": avg_days_per_team,
        "avg_days_per_stint": avg_days_per_stint,
        "pitcher_share": pitcher_share,
        "per_team_counts": counts,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Summarize MLB injury baselines.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="Path to the roster-resource injury workbook.")
    parser.add_argument("--json", action="store_true", help="Emit raw JSON instead of formatted text.")
    parser.add_argument(
        "--league-size",
        type=int,
        action="append",
        help="Optional league sizes to project injury totals for (can be repeated).",
    )
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    stats = summarize(load_injury_records(args.workbook))
    if args.json:
        print(json.dumps(stats, default=lambda o: list(o.items()) if isinstance(o, Counter) else o, indent=2))
        return

    print(f"Teams tracked: {stats['teams_tracked']}")
    print(f"Total injuries logged: {stats['total_records']}")
    print(f"Avg injuries per team: {stats['avg_injuries_per_team']:.2f}")
    print(f"Avg days lost per team: {stats['avg_days_lost_per_team']:.1f}")
    print(f"Avg days per stint (where dates known): {stats['avg_days_per_stint']:.1f}")
    print(f"Pitcher share of injuries: {stats['pitcher_share']:.1%}")
    print()
    print("Top 5 teams by recorded injuries:")
    for team, count in stats["per_team_counts"].most_common(5):
        print(f"  {team}: {count}")

    if args.league_size:
        print()
        print("Projected totals by league size:")
        seen = set()
        for size in args.league_size:
            if size in seen or size <= 0:
                continue
            seen.add(size)
            total_inj = stats["avg_injuries_per_team"] * size
            total_days = stats["avg_days_lost_per_team"] * size
            print(f"  {size} teams -> ~{total_inj:.1f} IL stints, {total_days:.0f} days lost")


if __name__ == "__main__":
    main()
